#Load workbook. Load the worksheet 'Staff Initials'
import openpyxl
wb = openpyxl.load_workbook('13 - 19 February 2022.xlsx')
staff_initials = wb['Staff Initials']

si = []
d = []

for rectangle in staff_initials['A2:A44']:
  for cell in rectangle:
    si.append(cell.value)
print(si)

for rectangle in staff_initials['D2:D44']:
  for cell in rectangle:
    d.append
print(d)

#Create empty list to insert values of cells
initials = []

#rowOfCellOb == giant rectangle from F7:T29
#For each thing (in this case cell) in rowOfCellOb, find the actual content of the cell (using .value) and add it to the empty list "initials" (using .append)
for sheet in wb.worksheets:
  for rowOfCellOb in sheet['F7':'T29']:
    for each in rowOfCellOb:
      #Avoid adding empty cells to list with "if". If the value of the cell is not None, then you can append it to the list.
      if each.value is not None:
        initials.append(each.value)
#To compensate for too much data on Sunday, Friday, and Saturday, re-populate the "initials" list. Take entry ("each") for each item in "initials" with a length less than or equal to 4. This leaves two, three, and four letter intials. 
initials = [each for each in initials if len(each) <= 4]

print(initials)

#Create an empty dictionary called "hours". We will make each item in "initials" a key, and the value will be a count of the amount of instances of that key.
hours = {}
for i in initials:
    if i in hours:
        hours[i] += 1
    else:
        hours[i] = 1


for each in si:
  if each in hours:
    pass
  else:
    hours.update({each:0})

a = hours.items()
sorted_hours = sorted(a)

print (sorted_hours)

for rowNum in range(2, staff_initials.max_row + 1):
  enter = staff_initials.cell(row=rowNum, column = 1).value
  if enter in hours:
    staff_initials.cell(row=rowNum, column=4).value = hours[enter]

wb.save('updatedSched 13 - 19 Feb.xlsx')
#Now want to take first column of "Staff Initials" worksheet, skip first row (it's a header) and match to "hours" keys, then write "hours" values in Column D.